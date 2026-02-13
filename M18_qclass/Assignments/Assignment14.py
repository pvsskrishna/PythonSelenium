"""
Day 17,18

assignment:
-----------
ws to launch https://www.naukri.com/registration/createAccount?othersrcp=23531&wExp=N&utm_source=google&utm_medium=cpc&utm_campaign=Brand&gclsrc=aw.ds&gad_source=1&gad_campaignid=19863995494&gbraid=0AAAAADLp3cEygg3qqw3KWN4HIBioCYA2e&gclid=Cj0KCQiAvOjKBhC9ARIsAFvz5lguz00zGqk13g89vp5afVI5YcUzCKh0x2PJ4vdFcMR8uaiJfCVUnoQaAq6tEALw_wcB
create 5 different naukri profile, read data from excel file.
"""
import time
from datetime import datetime

import openpyxl
from selenium.webdriver import Chrome
from selenium.webdriver.chrome.options import ChromiumOptions
from selenium.webdriver.common.by import By

o=ChromiumOptions()
o.add_argument("--disable-notifications")
o.add_experimental_option("detach", False)
driver = Chrome(options=o)

def takeScreenshot():
    screenshotName = datetime.now().strftime("%y-%m-%d_%H-%M-%S")
    driver.get_screenshot_as_file(f"./screenshots_test/{screenshotName}.png")

def testCase1():
    try:
        driver.get("https://www.irctc.co.in/nget/train-search")
        driver.maximize_window()
        #driver.find_element(By.XPATH,"")
        time.sleep(2)
    except Exception as e:
        takeScreenshot()
        print(f"Error is: {e}")
        raise
    finally:
        driver.quit()

from openpyxl import *

def testCase2():
    try:
        workbook1 = openpyxl.load_workbook(r"./datasets_excels/data.xlsx")
        print('opened')
        sheet = workbook1.active
        print(sheet['A1'].value)
    except Exception as e:
        takeScreenshot()
        print(f"Error is: {e}")
        raise

def testCase3():
    try:
        workbook1 = openpyxl.load_workbook(r"./datasets_excels/data.xlsx")
        sheet = workbook1['Sheet1']
        data = sheet[2]
        print([i.value for i in data])

        for col in range(3,5):
            print(sheet.cell(row=1,column=col).value)

    except Exception as e:
        takeScreenshot()
        print(f"Error is: {e}")
        raise

def testCase4():
    try:
        workbook1 = openpyxl.load_workbook(r"./datasets_excels/data.xlsx")
        sheet = workbook1["Sheet1"]
        maxrows = sheet.max_row
        maxcol = sheet.max_column
        print(maxrows)
        print(maxcol)

    except Exception as e:
        takeScreenshot()
        print(f"Error is: {e}")
        raise

def testCase5():
    try:
        workbook1 = openpyxl.load_workbook(r"./datasets_excels/data.xlsx")
        sheet = workbook1["Sheet1"]

        for col in range(1, sheet.max_column+1):
            if sheet.cell(row=1,column=col).value == 'username':
                usernameColNo = col
            elif sheet.cell(row=1,column=col).value == 'password':
                passwordColNo = col

        creds = {}
        for row in range(1, sheet.max_row+1):
            creds[sheet.cell(row= row,column = usernameColNo).value] = sheet.cell(row=row,column=passwordColNo).value
        print(creds,end="\n")

        for un,pwd in creds.items():
            driver.implicitly_wait(20)
            driver.get("https://www.facebook.com/")
            driver.maximize_window()
            driver.find_element("id", "email").send_keys(un)
            driver.find_element("id", "pass").send_keys(pwd)
            driver.find_element("name", "login").click()
            time.sleep(5)
        driver.quit()

    except Exception as e:
        takeScreenshot()
        print(f"Error is: {e}")
        raise
