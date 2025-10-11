import openpyxl

book = openpyxl.load_workbook("Demo.xlsx")
sheet = book.active
cell = sheet.cell(row=1,column=1)
print(cell.value)
#sheet.cell(row=2,column=2).value = "Varun"
# print(sheet.cell(row=2,column=2).value)
# print(sheet.max_row)
# print(sheet.max_column)
# print(sheet["A1"].value)

# hwo to print all the values present in the sheet
# we will be writing one simple for loop

rows = sheet.max_row
columns = sheet.max_column
for row in range(1,rows+1):
    for column in range(1,columns+1):
        print(sheet.cell(row=row, column=column).value)


#To print any desired row values
rows = sheet.max_row
columns = sheet.max_column
for row in range(1,rows+1):
    if sheet.cell(row=row, column=1).value == 'Testcase2':
        for column in range(1,columns+1):
            print(sheet.cell(row=row, column=column).value)

# To store our data inside a dictionary
Dict = {}
rows = sheet.max_row
columns = sheet.max_column
for row in range(1,rows+1):
    if sheet.cell(row=row, column=1).value == 'Testcase2':
        for column in range(1,columns+1):
            #{'Name': 'Testcase2', 'firstname': 'c', 'lastname': 'd', 'email': 'cd.com'}
            Dict[sheet.cell(row=1, column=column).value] = sheet.cell(row=row, column=column).value
print(Dict)
