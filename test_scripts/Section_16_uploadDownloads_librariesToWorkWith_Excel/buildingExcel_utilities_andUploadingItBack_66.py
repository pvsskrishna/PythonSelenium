# This is a end to end functional testing scenario we are constructing, along with upload and download.
# Scenario :: 1. download the Excel sheet.
# 2. change the values.
# 3. save it and upload the file.
# 4. Check the toaster message appear and disappear
# 5. also values updated or not in table.


import time as t
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common import action_chains
from selenium.webdriver.support import expected_conditions as EC
import openpyxl


def update_excel_data(filePath, itemName, columnName, newValue):
    # edit the Excel with updated values now --> we will be updating the excel
    # step1: 1st dynamically automatically find out the desired row ---> just like how we found the fruit name

    book = openpyxl.load_workbook(filePath)
    sheet = book.active
    rows = sheet.max_row
    columns = sheet.max_column
    Dict = {}
    for i in range(1, rows + 1):
        if sheet.cell(row=i, column=2).value == itemName:
            Dict["row"] = i

    # step2: after locking the row,we will be updating the value inside the desire column  ---> just like how we extracted price, we will be updating the value.
    for j in range(1, columns + 1):
        if sheet.cell(row=1, column=j).value == columnName:
            Dict["column"] = j

    # Basically we are finding row and column numbers to keep in cell(row=n,column=m)

    # updating the new value
    sheet.cell(row=Dict["row"], column=Dict["column"]).value = newValue

    # saving the book after
    book.save(filePath)

chrome_options = webdriver.ChromeOptions()
#chrome_options.add_argument('headless')
chrome_options.add_argument("--start-maximized")

service_obj = Service(r"C:\Users\varun\PycharmProjects\PythonSelenium\Drivers\chromedriver.exe")
#driver = webdriver.Chrome(service = service_obj,options=chrome_options)
driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=chrome_options)
driver.implicitly_wait(10)
wait = WebDriverWait(driver,20)
action = action_chains.ActionChains(driver)

fruit_name = 'Apple'
url = "https://rahulshettyacademy.com/upload-download-test/index.html"
file_path = r"C:\Users\varun\Downloads\download.xlsx"
newValue = '999'
driver.get(url)
t.sleep(2)

#downloading the document
driver.find_element(By.ID,"downloadButton").click()

#updating the data in file
update_excel_data(file_path, fruit_name, "price",newValue)

#Uploading the file
file_input = driver.find_element(By.CSS_SELECTOR,"input[type='file']")
file_input.send_keys(file_path)

#checking the toaster
toaster_locator = (By.CSS_SELECTOR,"div[class='Toastify__toast-body'] div:nth-child(2)")
wait.until(EC.visibility_of_element_located(toaster_locator))
print(driver.find_element(*toaster_locator).text) #here * unpacks the locator tuple (By, value) into two separate arguments for find_element.
t.sleep(5)

#extracting fruit actual price from the table
priceColumnNo = driver.find_element(By.XPATH,"//div[text()='Price']").get_attribute('data-column-id')
actual_price = driver.find_element(By.XPATH,"//div[text() ='"+fruit_name+"']/parent::div/parent::div//div[@id='cell-"+priceColumnNo+"-undefined']").text
print(actual_price)


assert actual_price == newValue

